"""Validation for the excel-report example (issue #755)."""

from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path
from typing import Any

import pytest

REPO_ROOT = Path(__file__).resolve().parents[3]
EXAMPLE_DIR = REPO_ROOT / "examples" / "excel-report"
EXAMPLE_PROCESS = EXAMPLE_DIR / "process.json"
SAMPLE_CSV = EXAMPLE_DIR / "sample" / "sales_data.csv"
REPORT_XLSX = EXAMPLE_DIR / "output" / "report.xlsx"

from rpaforge.core.diagram_converter import DiagramConverter  # noqa: E402


def _load_script_module():
    """Import examples/excel-report/script.py as a module."""
    spec = importlib.util.spec_from_file_location(
        "excel_report_script", EXAMPLE_DIR / "script.py"
    )
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class TestBundledSample:
    def test_sample_csv_exists_and_contains_expected_headers(self) -> None:
        assert SAMPLE_CSV.is_file()
        content = SAMPLE_CSV.read_text(encoding="utf-8")
        assert "date,region,category,product,units,unit_price,revenue" in content
        assert "Electronics" in content
        assert "Furniture" in content
        assert "Supplies" in content


class TestDiagramConversion:
    """Acceptance: process.json passes the DiagramConverter smoke-check."""

    def _diagram(self) -> dict[str, Any]:
        with open(EXAMPLE_PROCESS, encoding="utf-8") as fh:
            return dict(json.load(fh))

    def test_example_converts_to_expected_chain(self) -> None:
        diagram = self._diagram()
        process = DiagramConverter().convert(diagram)
        calls = [
            (a.library, a.activity) for task in process.tasks for a in task.activities
        ]
        assert ("DataFrames", "Read CSV") in calls
        assert ("DataFrames", "Group By") in calls
        assert ("DataFrames", "Sort") in calls
        assert ("DataFrames", "Write Excel") in calls

        read_idx = calls.index(("DataFrames", "Read CSV"))
        assert calls[read_idx + 1] == ("DataFrames", "Group By")
        assert calls[read_idx + 2] == ("DataFrames", "Sort")
        assert calls[read_idx + 3] == ("DataFrames", "Write Excel")

    def test_output_variables_chained(self) -> None:
        diagram = self._diagram()
        process = DiagramConverter().convert(diagram)
        outputs = {
            a.activity: a.output_variable
            for task in process.tasks
            for a in task.activities
        }
        assert outputs["Read CSV"] == "dfName"
        assert outputs["Group By"] == "summaryFrame"
        assert outputs["Sort"] == "sortedFrame"
        assert outputs["Write Excel"] == "savedPath"


class TestScriptEndToEnd:
    """Acceptance: script.py produces report.xlsx from the bundled CSV."""

    def test_script_produces_report_xlsx(self) -> None:
        pytest.importorskip("polars")
        pytest.importorskip("openpyxl")

        REPORT_XLSX.unlink(missing_ok=True)
        module = _load_script_module()

        exit_code = module.main()

        assert exit_code == 0
        assert REPORT_XLSX.is_file()

        import openpyxl

        wb = openpyxl.load_workbook(str(REPORT_XLSX))
        assert "Category Summary" in wb.sheetnames
        ws = wb["Category Summary"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("category", "revenue")
        # Ensure categories are sorted descending by revenue
        assert rows[1][0] == "Electronics"
        assert rows[1][1] == 9575
        assert rows[2][0] == "Furniture"
        assert rows[2][1] == 3600
        assert rows[3][0] == "Supplies"
        assert rows[3][1] == 490
